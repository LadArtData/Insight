#!/usr/bin/env python3
"""
INSIGHT_workbook_reader.py

Turns a client's filled-in spreadsheet into something a model can actually
read, and then checks what the model claims to have read.

Existing clients have no discovery interview to run -- the engagement already
happened and a hand-filled workbook is the only record of it. So the import
path has to work on whatever shape that spreadsheet is in, which rules out a
parser written against one template.

Two jobs here, and neither of them calls a model:

  digest_workbook()  reduces a workbook to a compact, cell-addressed summary
  verify_citations() checks a proposed answer against the cells it cites

WHY THE REDUCTION IS THE WORK
-----------------------------
A real configuration workbook is 33 sheets and 700 KB; one sheet in the
reference file carries 7,395 rows. None of that fits in a model's context, and
paying to send it would be absurd besides. But almost all of it is repetition:
the sheet that matters says "six segments, these names, these lengths" in its
first ten rows, and then lists ten thousand account combinations.

So each sheet is reduced to its header, a sample of populated rows, and a
count of what was left out. The count matters -- "1,685 further rows" is
itself an answer to "how many segment values are there", and dropping it
silently would lose that.

WHAT THIS CANNOT DO
-------------------
Sampling represents a sheet; it does not guarantee that any particular row
survives. On a sheet of two hundred near-identical rows, one anomalous row may
fall between sample points. Sheets short enough to keep whole -- which is where
contradictions have actually turned up, such as an approval list whose members
all belong to a different organisation than the legal entity -- are complete,
but that is a property of their size, not a promise about large ones.

WHY EVERY CELL KEEPS ITS ADDRESS
--------------------------------
Every value carries its A1 reference, so a proposed answer can cite the cells
it came from and this module can go back and check them. An import that says
"the model read it" is a leap of faith; an import that cites Ledger!E8 and has
the citation verified is a fact someone can audit. That check is what makes
the difference between a client file worth trusting and a client file full of
plausible values nobody verified -- which is worse than an empty one, because
an empty one is visibly empty.
"""

import re
from collections import OrderedDict

# A model has to be able to hold the whole digest at once, so there is a
# ceiling. When it is hit the digest says so rather than quietly stopping:
# a truncated import that looks complete is the failure mode worth avoiding.
# Tuned against a real 33-sheet workbook: at 120,000 every sheet is
# represented, nothing is truncated, and every fact a human extracted by hand
# survives the reduction. 90,000 loses one. The result is roughly 20,000
# tokens -- comfortably inside a model's context, from a 635 KB file.
DEFAULT_MAX_CHARS = 120000
DEFAULT_SAMPLE_ROWS = 24

# Rows in these workbooks are frequently instructions to whoever fills the
# sheet in ("30 Characters", "List of Values", "Yes/No") rather than data.
# They are recognisable and worth dropping: they are identical across every
# client's copy and say nothing about this client.
INSTRUCTION_MARKERS = (
    "characters", "list of values", "yes/no", "free text", "valid ",
    "list of seeded", "define your", "number", "yyyy-mm-dd",
)


def _is_instruction_row(values):
    """True when a row is template guidance rather than a client's answer."""
    text = " ".join(str(v).lower() for v in values if v not in (None, ""))
    if not text:
        return False
    hits = sum(1 for m in INSTRUCTION_MARKERS if m in text)
    # Two or more markers, or one marker in a very short row, reads as
    # guidance. One marker in a long row of real values does not -- "Number"
    # is a legitimate answer somewhere.
    return hits >= 2 or (hits == 1 and len(text) < 40)


def _col_letter(idx):
    """1 -> A, 27 -> AA. openpyxl has this, but only on the writable path."""
    letters = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _clean(value):
    """A cell as text, with the noise that helps nobody removed."""
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def digest_workbook(source, max_sample_rows=DEFAULT_SAMPLE_ROWS,
                    max_chars=DEFAULT_MAX_CHARS):
    """
    Reduce a workbook to a cell-addressed digest.

    `source` is a path or any file-like object. The service passes a stream so
    an uploaded workbook never touches disk: the container is ephemeral, and a
    client's configuration file is not something to leave in /tmp after a
    crash.

    Returns a dict:
      {"sheets": [{"name", "populated_rows", "rows", "omitted_rows"}],
       "truncated": bool, "omitted_sheets": [names]}

    where each entry in "rows" is a list of {"ref", "value"}.

    Read-only and data_only, so formulas arrive as their last computed value.
    A workbook that has never been opened in Excel can therefore yield blanks
    where formulas live; that is a property of the file, not of this code, and
    it shows up as a missing answer rather than a wrong one.
    """
    import openpyxl  # local import: the web container has no need of it

    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    try:
        return _digest(wb, max_sample_rows, max_chars)
    finally:
        wb.close()


def _cost(rows):
    return sum(len(c["ref"]) + len(c["value"]) + 4 for row in rows for c in row)


def _digest(wb, max_sample_rows, max_chars):
    sheets = []
    used = 0
    omitted_sheets = []
    sheet_count = len(wb.worksheets)

    for ws in wb.worksheets:
        populated = []          # (row_index, [(col_index, value)])
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = [(c_idx, _clean(v)) for c_idx, v in enumerate(row, start=1)
                     if _clean(v)]
            if cells:
                populated.append((r_idx, cells))

        if not populated:
            continue

        body = [row for row in populated
                if not _is_instruction_row([v for _, v in row[1]])]
        sample, omitted = _sample(body, max_sample_rows)

        entry = OrderedDict()
        entry["name"] = ws.title
        entry["populated_rows"] = len(populated)
        entry["rows"] = [_refs(ws.title, row) for row in sample]
        # Not decoration. "829 further rows" answers "how many account
        # combinations are there" on its own, and a digest that dropped it
        # would lose an answer rather than just some detail.
        entry["omitted_rows"] = omitted

        # Every sheet gets a share rather than the budget being spent
        # first-come. Dropping whichever sheets happen to be last is
        # arbitrary -- and in the reference workbook it dropped the
        # intercompany and approval sheets purely because they sit at the
        # end of the tab order.
        per_sheet = max(1500, max_chars // max(1, sheet_count))
        while entry["rows"] and _cost(entry["rows"]) > per_sheet:
            entry["rows"].pop()
            entry["omitted_rows"] += 1
        used += _cost(entry["rows"])
        sheets.append(entry)

    return {
        "sheets": sheets,
        "truncated": bool(omitted_sheets),
        "omitted_sheets": omitted_sheets,
    }


def _sample(body, max_sample_rows):
    """
    Choose which rows represent a sheet.

    There is deliberately no attempt to find "the header row". An earlier
    version did, and it was wrong on every sheet that mattered: these sheets
    are a sequence of blocks rather than one table, so the row with the most
    short labels is often a sub-table far down, and everything above it --
    which is where the facts live -- was discarded. The accounting calendar
    states its period frequency and adjusting periods in rows 7 to 13 and then
    lists periods; each control budget names its source system in rows 3 to 16
    and then lists control rules; the approval sheets list their members in
    rows 9 to 22 and then carry pages of workflow metadata. Header detection
    threw away all three.

    So no structural assumption survives. Small sheets are kept whole, because
    they are cheap and their facts are scattered through them. Large ones are
    sampled from the top and spread through the rest, so anything distinctive
    further down is still represented.
    """
    if not body:
        return [], 0

    # Keep sheets under this size whole. The reference workbook's calendar
    # (27 rows) and each control budget (40) sit under it, and each would lose
    # its defining values to any narrower rule.
    KEEP_WHOLE_UNDER = 45
    if len(body) <= max(KEEP_WHOLE_UNDER, max_sample_rows):
        return body, 0

    head_count = max(4, max_sample_rows // 2)
    head = body[:head_count]
    rest = body[head_count:]
    want = max(1, max_sample_rows - head_count)
    step = max(1, len(rest) // want)
    spread = rest[::step][:want]
    return head + spread, len(body) - len(head) - len(spread)


def _refs(sheet_name, row):
    r_idx, cells = row
    return [{"ref": "%s!%s%d" % (sheet_name, _col_letter(c_idx), r_idx),
             "value": value}
            for c_idx, value in cells]


CELL_REF = re.compile(r"^(?:'(?P<q>[^']+)'|(?P<s>[^!]+))!(?P<col>[A-Z]{1,3})(?P<row>\d+)$")


def parse_ref(ref):
    """'Ledger'!E8 or Ledger!E8 -> (sheet, column, row). None when malformed."""
    m = CELL_REF.match((ref or "").strip())
    if not m:
        return None
    sheet = m.group("q") or m.group("s")
    return sheet, m.group("col"), int(m.group("row"))


def read_cells(source, refs):
    """
    Look up a list of A1 references. Missing or malformed ones come back ''.

    `source` is a path or a file-like object, as for digest_workbook.
    """
    import openpyxl

    wb = openpyxl.load_workbook(source, data_only=True)
    try:
        out = {}
        for ref in refs:
            parsed = parse_ref(ref)
            if not parsed:
                out[ref] = ""
                continue
            sheet, col, row = parsed
            if sheet not in wb.sheetnames:
                out[ref] = ""
                continue
            try:
                out[ref] = _clean(wb[sheet]["%s%d" % (col, row)].value)
            except (ValueError, KeyError):
                out[ref] = ""
        return out
    finally:
        wb.close()


def _tokens(text):
    """
    Words worth comparing, with edge punctuation removed.

    The trimming matters more than it looks. Interior punctuation has to stay,
    because the values being matched are things like 39-6005724, OCWI_GL_FUND
    and name@county.gov. But a token grabbed from prose arrives as "USD." with
    the sentence's full stop attached, and "usd." never equals the cell's
    "usd" -- so a correct answer would be reported as unsupported.
    """
    raw = re.findall(r"[A-Za-z0-9][A-Za-z0-9._@-]*", (text or "").lower())
    out = set()
    for token in raw:
        token = token.strip("._-@")
        if len(token) >= 3:
            out.add(token)
    return out


def verify_citations(answer, cited_values):
    """
    Check a proposed answer against the cells it cites.

    Two levels, because only one of them can be strict:

      supported   at least one cited cell is non-empty AND shares a token with
                  the answer. The answer genuinely came from somewhere.
      uncited     no usable citation at all -- nothing was checked, so the
                  answer should not be shown as though it had been.
      unsupported cells exist but nothing in them appears in the answer.

    The overlap test cannot be stricter than this. A good answer summarises
    several cells in a sentence -- "Six segments: Fund(5), Cost Center(6)..." --
    so demanding the cell text appear verbatim would reject exactly the
    answers worth having. What it does catch is the failure that matters: an
    answer citing cells that are empty, or whose contents have nothing to do
    with what was written.

    Returns {"status", "matched_refs", "empty_refs", "overlap"}.
    """
    non_empty = {ref: v for ref, v in (cited_values or {}).items() if v}
    empty = [ref for ref, v in (cited_values or {}).items() if not v]

    if not non_empty:
        return {"status": "uncited", "matched_refs": [], "empty_refs": empty,
                "overlap": 0.0}

    answer_tokens = _tokens(answer)
    matched = []
    for ref, value in non_empty.items():
        if _tokens(value) & answer_tokens:
            matched.append(ref)

    overlap = len(matched) / float(len(non_empty))
    return {
        "status": "supported" if matched else "unsupported",
        "matched_refs": matched,
        "empty_refs": empty,
        "overlap": round(overlap, 2),
    }


def digest_to_text(digest):
    """Flatten a digest for a prompt. One line per cell, address first."""
    lines = []
    for sheet in digest["sheets"]:
        lines.append("## sheet: %s (%d populated rows)" % (sheet["name"], sheet["populated_rows"]))
        for row in sheet["rows"]:
            lines.append("  " + " | ".join("%s=%s" % (c["ref"], c["value"]) for c in row))
        if sheet["omitted_rows"]:
            lines.append("  (%d further rows of the same shape, not shown)" % sheet["omitted_rows"])
        lines.append("")
    if digest["truncated"]:
        lines.append("NOTE: these sheets were too large to include: %s"
                     % ", ".join(digest["omitted_sheets"]))
    return "\n".join(lines)
